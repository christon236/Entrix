import datetime
from datetime import timedelta
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views import View

from main_app.models import Attendance, AttendanceVisit, Member, TrainerAttendance, TrainerAttendanceVisit
from masters.models import Trainer
from transactions.models import AttendanceLog
from transactions.services import build_member_detail_payload, build_trainer_detail_payload
from .forms import AttendanceReportFilterForm


class AttendanceReportView(LoginRequiredMixin, View):
    """
    Reports module view.
    Provides attendance analytics with member/trainer tabs,
    filtering by preset/custom date range, member search, and plan.
    Handles AJAX requests for member/trainer detail popups and record deletion.
    """

    template_name = "reports/attendance_report.html"

    def get(self, request, *args, **kwargs):
        # Handle AJAX requests
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            action = request.GET.get("action")
            if action == "get_member_details":
                return self._ajax_member_details(request)
            elif action == "get_trainer_details":
                return self._ajax_trainer_details(request)
            elif action == "get_member_logs":
                return self._ajax_member_logs(request)

        # Export the currently-filtered records to a formatted Excel workbook.
        if request.GET.get("export") == "excel":
            return self._export_excel(request)

        context = self._build_context(request)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        if action == "delete_attendance":
            return self._handle_delete_attendance(request)
        elif action == "delete_trainer_attendance":
            return self._handle_delete_trainer_attendance(request)
        messages.error(request, "Unknown action.")
        return redirect("attendance-report")

    # ---------------------------------------------------------------
    # Context Builder
    # ---------------------------------------------------------------

    def _resolve_filters(self, request, form):
        """Resolve the date range, search keyword, plan and active tab from the
        request's filter form. Shared by the HTML view and the Excel export so
        both always operate on exactly the same set of records."""
        today = timezone.localdate()
        start_date = today.replace(day=1)
        end_date = today
        date_preset = "this_month"
        member_search = ""
        plan_filter = None
        active_tab = request.GET.get("tab", "members")

        if form.is_valid():
            date_preset = form.cleaned_data.get("date_preset") or "this_month"
            member_search = form.cleaned_data.get("member_search") or ""
            plan_filter = form.cleaned_data.get("membership_plan")

            if date_preset == "today":
                start_date = today
                end_date = today
            elif date_preset == "yesterday":
                start_date = today - timedelta(days=1)
                end_date = today - timedelta(days=1)
            elif date_preset == "last_7_days":
                start_date = today - timedelta(days=6)
                end_date = today
            elif date_preset == "this_month":
                start_date = today.replace(day=1)
                end_date = today
            elif date_preset == "last_month":
                first_day_this_month = today.replace(day=1)
                last_day_last_month = first_day_this_month - timedelta(days=1)
                start_date = last_day_last_month.replace(day=1)
                end_date = last_day_last_month
            elif date_preset == "custom":
                s_val = form.cleaned_data.get("start_date")
                e_val = form.cleaned_data.get("end_date")
                if s_val and e_val:
                    start_date = s_val
                    end_date = e_val
                elif s_val:
                    start_date = s_val
                    end_date = s_val
                elif e_val:
                    start_date = e_val
                    end_date = e_val

        return start_date, end_date, date_preset, member_search, plan_filter, active_tab

    def _build_context(self, request):
        form = AttendanceReportFilterForm(request.GET or None)
        (
            start_date,
            end_date,
            date_preset,
            member_search,
            plan_filter,
            active_tab,
        ) = self._resolve_filters(request, form)

        today = timezone.localdate()
        form.initial["start_date"] = start_date.strftime("%Y-%m-%d")
        form.initial["end_date"] = end_date.strftime("%Y-%m-%d")

        # ---- Dashboard Cards (Change 2) ----
        total_report_dates = Attendance.objects.filter(
            date__range=[start_date, end_date]
        ).values("date").distinct().count()

        active_memberships_count = Member.objects.filter(
            is_active=True, membership_end_date__gte=today
        ).count()

        expired_memberships_count = Member.objects.filter(
            membership_end_date__lt=today
        ).count()

        # ---- Build Records (Members or Trainers tab) ----
        date_groups = []
        records_count = 0

        if active_tab == "trainers":
            date_groups, records_count = self._build_trainer_groups(
                start_date, end_date, member_search
            )
        else:
            date_groups, records_count = self._build_member_groups(
                start_date, end_date, member_search, plan_filter
            )

        page_number = request.GET.get("page", 1)
        date_groups_page = Paginator(date_groups, 10).get_page(page_number)

        return {
            "filter_form": form,
            "start_date": start_date,
            "end_date": end_date,
            "date_preset": date_preset,
            "total_report_dates": total_report_dates,
            "active_memberships_count": active_memberships_count,
            "expired_memberships_count": expired_memberships_count,
            "date_groups": date_groups_page,
            "records_count": records_count,
            "active_tab": active_tab,
            "has_filters": bool(member_search or plan_filter or date_preset != "this_month"),
        }

    def _build_member_groups(self, start_date, end_date, member_search, plan_filter):
        today = timezone.localdate()
        qs = Attendance.objects.select_related(
            "member", "member__membership_plan"
        ).prefetch_related("visits").filter(date__range=[start_date, end_date])

        if member_search:
            qs = qs.filter(
                Q(member__member_id__icontains=member_search) |
                Q(member__full_name__icontains=member_search) |
                Q(member__mobile_number__icontains=member_search) |
                Q(fingerprint_id__icontains=member_search)
            )

        if plan_filter:
            qs = qs.filter(member__membership_plan=plan_filter)

        raw_list = list(qs.order_by("-date", "-entry_time", "-id"))
        grouped = {}
        for att in raw_list:
            key = (att.member_id, att.date)
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(att)

        dedup_list = []
        for key, group in grouped.items():
            primary_att = group[0]
            all_visits = []
            for att in group:
                vs = list(att.visits.all())
                if not vs and att.entry_time:
                    vs = [AttendanceVisit(attendance=att, visit_number=1, entry_time=att.entry_time, exit_time=att.exit_time)]
                all_visits.extend(vs)

            all_visits.sort(key=lambda v: (v.entry_time or datetime.time.min, v.visit_number))
            for idx, v in enumerate(all_visits, start=1):
                if v.visit_number != idx:
                    v.visit_number = idx

            primary_att.visit_list = all_visits
            primary_att.visits_count = len(all_visits)
            if len(all_visits) <= 1:
                primary_att.visit_label = ""
            elif len(all_visits) == 2:
                primary_att.visit_label = "2nd Time"
            elif len(all_visits) == 3:
                primary_att.visit_label = "3rd Time"
            else:
                primary_att.visit_label = f"{len(all_visits)}th Time"

            total_seconds = 0
            if all_visits:
                primary_att.entry_time = all_visits[0].entry_time or primary_att.entry_time
                primary_att.exit_time = all_visits[-1].exit_time if all(v.exit_time is not None for v in all_visits) else None
                for v in all_visits:
                    if v.entry_time and v.exit_time:
                        entry_dt = datetime.datetime.combine(primary_att.date, v.entry_time)
                        exit_dt = datetime.datetime.combine(primary_att.date, v.exit_time)
                        if exit_dt < entry_dt:
                            exit_dt += timedelta(days=1)
                        sec = (exit_dt - entry_dt).total_seconds()
                        if sec > 0:
                            total_seconds += sec
                    elif v.entry_time and not v.exit_time and primary_att.date == today:
                        # Live, still-inside visit: accrue duration up to "now"
                        # exactly like Attendance Management (single source of truth).
                        entry_dt = datetime.datetime.combine(primary_att.date, v.entry_time)
                        now_dt = datetime.datetime.combine(today, timezone.localtime().time())
                        sec = (now_dt - entry_dt).total_seconds()
                        if sec > 0:
                            total_seconds += sec
            elif primary_att.entry_time and primary_att.exit_time:
                entry_dt = datetime.datetime.combine(primary_att.date, primary_att.entry_time)
                exit_dt = datetime.datetime.combine(primary_att.date, primary_att.exit_time)
                if exit_dt < entry_dt:
                    exit_dt += timedelta(days=1)
                sec = (exit_dt - entry_dt).total_seconds()
                if sec > 0:
                    total_seconds += sec
            elif primary_att.entry_time and not primary_att.exit_time and primary_att.date == today:
                entry_dt = datetime.datetime.combine(primary_att.date, primary_att.entry_time)
                now_dt = datetime.datetime.combine(today, timezone.localtime().time())
                sec = (now_dt - entry_dt).total_seconds()
                if sec > 0:
                    total_seconds += sec

            if total_seconds > 0:
                hours, remainder = divmod(int(total_seconds), 3600)
                minutes, _ = divmod(remainder, 60)
                if hours > 0 and minutes > 0:
                    dur_str = f"{hours}h {minutes}m"
                elif hours > 0:
                    dur_str = f"{hours}h 00m"
                else:
                    dur_str = f"{max(1, minutes)}m"

                limit_mins = 24 * 60
                if primary_att.member and primary_att.member.membership_plan and primary_att.member.membership_plan.daily_access_hours and not primary_att.member.membership_plan.is_full_day_access:
                    limit_mins = primary_att.member.membership_plan.daily_access_hours * 60

                total_mins = int(total_seconds / 60)
                remaining_mins = limit_mins - total_mins

                if remaining_mins <= 0:
                    primary_att.duration_badge_class = "badge bg-danger-subtle text-danger border border-danger-subtle px-2 py-1 fw-semibold"
                    primary_att.total_day_duration_str = f"🔴 {dur_str}"
                elif remaining_mins <= 30:
                    primary_att.duration_badge_class = "badge bg-warning-subtle text-warning-emphasis border border-warning-subtle px-2 py-1 fw-semibold"
                    primary_att.total_day_duration_str = f"🟠 {dur_str}"
                else:
                    primary_att.duration_badge_class = "badge bg-success-subtle text-success border border-success-subtle px-2 py-1 fw-semibold"
                    primary_att.total_day_duration_str = f"🟢 {dur_str}"
            else:
                primary_att.total_day_duration_str = "--"
                primary_att.duration_badge_class = "badge bg-light text-dark border px-2 py-1 fw-semibold"

            dedup_list.append(primary_att)

        records_count = len(dedup_list)
        records_by_date = {}
        for att in dedup_list:
            if att.date not in records_by_date:
                records_by_date[att.date] = []
            records_by_date[att.date].append(att)

        date_groups = []
        for d_key, recs in records_by_date.items():
            d_mins = 0
            d_count = 0
            for r in recs:
                if r.visit_list:
                    for v in r.visit_list:
                        if v.entry_time and v.exit_time:
                            edt = datetime.datetime.combine(r.date, v.entry_time)
                            xdt = datetime.datetime.combine(r.date, v.exit_time)
                            if xdt < edt:
                                xdt += timedelta(days=1)
                            d_mins += (xdt - edt).total_seconds() / 60.0
                            d_count += 1
                elif r.entry_time and r.exit_time:
                    edt = datetime.datetime.combine(r.date, r.entry_time)
                    xdt = datetime.datetime.combine(r.date, r.exit_time)
                    if xdt < edt:
                        xdt += timedelta(days=1)
                    d_mins += (xdt - edt).total_seconds() / 60.0
                    d_count += 1
            d_avg = f"{int(d_mins // d_count // 60)}h {int((d_mins // d_count) % 60)}m" if d_count > 0 else "--"

            date_groups.append({
                "date": d_key,
                "date_formatted": d_key.strftime("%A, %d %B %Y"),
                "records": recs,
                "total_checkins": len(recs),
                "avg_duration": d_avg,
            })

        return date_groups, records_count

    def _build_trainer_groups(self, start_date, end_date, search):
        today = timezone.localdate()
        qs = TrainerAttendance.objects.select_related(
            "trainer"
        ).prefetch_related("visits").filter(date__range=[start_date, end_date])

        if search:
            qs = qs.filter(
                Q(trainer__trainer_id__icontains=search) |
                Q(trainer__full_name__icontains=search) |
                Q(trainer__mobile_number__icontains=search)
            )

        raw_list = list(qs.order_by("-date", "-entry_time", "-id"))
        grouped = {}
        for att in raw_list:
            key = (att.trainer_id, att.date)
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(att)

        dedup_list = []
        for key, group in grouped.items():
            primary_att = group[0]
            all_visits = []
            for att in group:
                vs = list(att.visits.all())
                if not vs and att.entry_time:
                    vs = [TrainerAttendanceVisit(attendance=att, visit_number=1, entry_time=att.entry_time, exit_time=att.exit_time)]
                all_visits.extend(vs)

            all_visits.sort(key=lambda v: (v.entry_time or datetime.time.min, v.visit_number))
            for idx, v in enumerate(all_visits, start=1):
                if v.visit_number != idx:
                    v.visit_number = idx

            primary_att.visit_list = all_visits
            primary_att.visits_count = len(all_visits)
            if len(all_visits) <= 1:
                primary_att.visit_label = ""
            elif len(all_visits) == 2:
                primary_att.visit_label = "2nd Time"
            elif len(all_visits) == 3:
                primary_att.visit_label = "3rd Time"
            else:
                primary_att.visit_label = f"{len(all_visits)}th Time"

            total_seconds = 0
            if all_visits:
                primary_att.entry_time = all_visits[0].entry_time or primary_att.entry_time
                primary_att.exit_time = all_visits[-1].exit_time if all(v.exit_time is not None for v in all_visits) else None
                for v in all_visits:
                    if v.entry_time and v.exit_time:
                        entry_dt = datetime.datetime.combine(primary_att.date, v.entry_time)
                        exit_dt = datetime.datetime.combine(primary_att.date, v.exit_time)
                        if exit_dt < entry_dt:
                            exit_dt += timedelta(days=1)
                        sec = (exit_dt - entry_dt).total_seconds()
                        if sec > 0:
                            total_seconds += sec
                    elif v.entry_time and not v.exit_time and primary_att.date == today:
                        # Live, still-inside visit: accrue duration up to "now"
                        # exactly like Attendance Management (single source of truth).
                        entry_dt = datetime.datetime.combine(primary_att.date, v.entry_time)
                        now_dt = datetime.datetime.combine(today, timezone.localtime().time())
                        sec = (now_dt - entry_dt).total_seconds()
                        if sec > 0:
                            total_seconds += sec
            elif primary_att.entry_time and primary_att.exit_time:
                entry_dt = datetime.datetime.combine(primary_att.date, primary_att.entry_time)
                exit_dt = datetime.datetime.combine(primary_att.date, primary_att.exit_time)
                if exit_dt < entry_dt:
                    exit_dt += timedelta(days=1)
                sec = (exit_dt - entry_dt).total_seconds()
                if sec > 0:
                    total_seconds += sec
            elif primary_att.entry_time and not primary_att.exit_time and primary_att.date == today:
                entry_dt = datetime.datetime.combine(primary_att.date, primary_att.entry_time)
                now_dt = datetime.datetime.combine(today, timezone.localtime().time())
                sec = (now_dt - entry_dt).total_seconds()
                if sec > 0:
                    total_seconds += sec

            if total_seconds > 0:
                hours, remainder = divmod(int(total_seconds), 3600)
                minutes, _ = divmod(remainder, 60)
                if hours > 0 and minutes > 0:
                    dur_str = f"{hours}h {minutes}m"
                elif hours > 0:
                    dur_str = f"{hours}h 00m"
                else:
                    dur_str = f"{max(1, minutes)}m"
            else:
                dur_str = "--"

            primary_att.total_day_duration_str = dur_str
            dedup_list.append(primary_att)

        records_count = len(dedup_list)
        records_by_date = {}
        for att in dedup_list:
            if att.date not in records_by_date:
                records_by_date[att.date] = []
            records_by_date[att.date].append(att)

        date_groups = []
        for d_key, recs in records_by_date.items():
            d_mins = 0
            d_count = 0
            for r in recs:
                if r.visit_list:
                    for v in r.visit_list:
                        if v.entry_time and v.exit_time:
                            edt = datetime.datetime.combine(r.date, v.entry_time)
                            xdt = datetime.datetime.combine(r.date, v.exit_time)
                            if xdt < edt:
                                xdt += timedelta(days=1)
                            d_mins += (xdt - edt).total_seconds() / 60.0
                            d_count += 1
                elif r.entry_time and r.exit_time:
                    edt = datetime.datetime.combine(r.date, r.entry_time)
                    xdt = datetime.datetime.combine(r.date, r.exit_time)
                    if xdt < edt:
                        xdt += timedelta(days=1)
                    d_mins += (xdt - edt).total_seconds() / 60.0
                    d_count += 1
            d_avg = f"{int(d_mins // d_count // 60)}h {int((d_mins // d_count) % 60)}m" if d_count > 0 else "--"

            date_groups.append({
                "date": d_key,
                "date_formatted": d_key.strftime("%A, %d %B %Y"),
                "records": recs,
                "total_checkins": len(recs),
                "avg_duration": d_avg,
            })

        return date_groups, records_count

    # ---------------------------------------------------------------
    # AJAX Handlers
    # ---------------------------------------------------------------

    def _ajax_member_details(self, request):
        """Return member profile + attendance summary.

        Delegates to the shared ``build_member_detail_payload`` so the Reports
        detail popup renders exactly the same fields (including Date of Birth,
        Email, Username and Address) as the Attendance Management and Dashboard
        popups — one single source of truth.
        """
        member_id = request.GET.get("member_id", "").strip()
        member = get_object_or_404(Member, member_id=member_id)
        return JsonResponse(build_member_detail_payload(member))

    def _ajax_trainer_details(self, request):
        """Return trainer profile details via the shared single-source payload."""
        trainer_id = request.GET.get("trainer_id", "").strip()
        trainer = get_object_or_404(Trainer, trainer_id=trainer_id)
        return JsonResponse(build_trainer_detail_payload(trainer))

    def _ajax_member_logs(self, request):
        """Return access audit trail logs for a member (existing functionality)."""
        member_id = request.GET.get("member_id", "").strip()
        if not member_id:
            return JsonResponse({"error": "Member ID is required."}, status=400)

        member = Member.objects.filter(member_id=member_id).first()
        if not member:
            return JsonResponse({"error": "Member not found."}, status=404)

        logs = AttendanceLog.objects.filter(member=member).order_by("-timestamp")[:20]
        log_list = []
        for l in logs:
            log_list.append({
                "id": l.pk,
                "event_type": l.get_event_type_display(),
                "timestamp": l.timestamp.strftime("%d %b %Y, %H:%M"),
                "entry_allowed": l.entry_allowed,
                "reason": l.reason or ("Access Granted" if l.entry_allowed else "Access Denied"),
                "fingerprint_id": l.fingerprint_id or member.fingerprint_id or "--",
            })

        history_qs = Attendance.objects.filter(member=member)
        total_visits = history_qs.count()

        return JsonResponse({
            "member_name": member.full_name,
            "member_id": member.member_id,
            "plan_name": member.membership_plan.name if member.membership_plan else "General Plan",
            "photo_url": member.photo.url if member.photo else f"https://ui-avatars.com/api/?name={member.full_name}&background=random&size=84",
            "total_visits": total_visits,
            "logs": log_list,
        })

    # ---------------------------------------------------------------
    # POST Handlers
    # ---------------------------------------------------------------

    def _handle_delete_attendance(self, request):
        """Delete a member attendance record."""
        att_id = request.POST.get("attendance_id")
        try:
            att = Attendance.objects.get(pk=att_id)
            member_name = att.member.full_name
            att_date = att.date.strftime("%d %b %Y")
            att.delete()
            messages.success(request, f"Attendance record for {member_name} on {att_date} deleted successfully.")
        except Attendance.DoesNotExist:
            messages.error(request, "Attendance record not found.")
        return redirect("attendance-report")

    def _handle_delete_trainer_attendance(self, request):
        """Delete a trainer attendance record."""
        att_id = request.POST.get("attendance_id")
        try:
            att = TrainerAttendance.objects.get(pk=att_id)
            trainer_name = att.trainer.full_name
            att_date = att.date.strftime("%d %b %Y")
            att.delete()
            messages.success(request, f"Trainer attendance record for {trainer_name} on {att_date} deleted successfully.")
        except TrainerAttendance.DoesNotExist:
            messages.error(request, "Trainer attendance record not found.")
        return redirect("attendance-report")

    # ---------------------------------------------------------------
    # Excel Export
    # ---------------------------------------------------------------

    def _visit_times_str(self, att):
        """Return a human-readable "HH:MM → HH:MM" list of all visits for the
        day, matching the Entry/Exit column shown in the report table."""
        pairs = []
        visits = getattr(att, "visit_list", None) or []
        if visits:
            for v in visits:
                entry = v.entry_time.strftime("%H:%M") if v.entry_time else "--"
                exit_ = v.exit_time.strftime("%H:%M") if v.exit_time else "--"
                pairs.append(f"{entry} → {exit_}")
        elif att.entry_time:
            entry = att.entry_time.strftime("%H:%M")
            exit_ = att.exit_time.strftime("%H:%M") if att.exit_time else "--"
            pairs.append(f"{entry} → {exit_}")
        return "\n".join(pairs) if pairs else "--"

    @staticmethod
    def _clean_duration(value):
        """Strip the coloured status emoji prefix from a duration string so the
        exported cell holds a clean "1h 30m" value."""
        if not value:
            return "--"
        for emoji in ("\U0001F534", "\U0001F7E0", "\U0001F7E2"):
            value = value.replace(emoji, "")
        return value.strip() or "--"

    def _export_excel(self, request):
        """Stream the currently-filtered member/trainer records as a formatted
        .xlsx workbook. Honours the active tab, date range and search filters so
        the export always mirrors exactly what the report table displays."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # openpyxl is an optional dependency (see requirements.txt). If it is
            # not installed in the environment, fail gracefully with a clear
            # message instead of raising a 500 ModuleNotFoundError.
            messages.error(
                request,
                "Excel export is unavailable because the 'openpyxl' package is "
                "not installed. Run 'pip install -r requirements.txt' to enable it.",
            )
            tab = "trainers" if request.GET.get("tab") == "trainers" else "members"
            return redirect(f"{reverse('attendance-report')}?tab={tab}")

        form = AttendanceReportFilterForm(request.GET or None)
        (
            start_date,
            end_date,
            date_preset,
            member_search,
            plan_filter,
            active_tab,
        ) = self._resolve_filters(request, form)

        is_trainers = active_tab == "trainers"
        if is_trainers:
            date_groups, _ = self._build_trainer_groups(start_date, end_date, member_search)
            headers = ["Date", "Trainer ID", "Trainer Name", "Mobile", "Designation",
                       "Gender / Age", "Email", "Joined Date",
                       "Entry / Exit Time", "Total Duration", "Visits"]
            sheet_title = "Trainer Attendance"
        else:
            date_groups, _ = self._build_member_groups(start_date, end_date, member_search, plan_filter)
            headers = ["Date", "Member ID", "Member Name", "Mobile", "Plan",
                       "Gender / Age", "Email", "Plan Joined Date", "Plan Expiry Date",
                       "Entry / Exit Time", "Total Duration", "Visits"]
            sheet_title = "Member Attendance"

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

        # ---- Styles ----
        brand_fill = PatternFill(start_color="2E6DA4", end_color="2E6DA4", fill_type="solid")
        title_font = Font(name="Calibri", size=14, bold=True, color="2E6DA4")
        sub_font = Font(name="Calibri", size=10, italic=True, color="666666")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Calibri", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D0D7DE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ---- Title block ----
        last_col = get_column_letter(len(headers))
        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = f"ENTRIX — {sheet_title} Report"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(f"A2:{last_col}2")
        ws["A2"] = (
            f"Range: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
            f"    |    Generated: {timezone.localtime().strftime('%d %b %Y, %H:%M')}"
        )
        ws["A2"].font = sub_font
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

        # ---- Header row ----
        header_row = 4
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.fill = brand_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        # ---- Data rows ----
        row = header_row + 1
        total_records = 0
        entry_exit_col = 9 if is_trainers else 10
        for dg in date_groups:
            for att in dg["records"]:
                if is_trainers:
                    t = att.trainer
                    gender_age = t.get_gender_display() if hasattr(t, "get_gender_display") else (t.gender or "")
                    if t.age is not None:
                        gender_age = f"{gender_age or '--'} / {t.age} yrs"
                    else:
                        gender_age = gender_age or "--"
                    values = [
                        att.date.strftime("%d %b %Y"),
                        t.trainer_id,
                        t.full_name,
                        t.mobile_number or "--",
                        (t.get_designation_display()
                         if hasattr(t, "get_designation_display")
                         else t.designation) or "--",
                        gender_age,
                        t.email or "--",
                        t.joining_date.strftime("%d %b %Y") if t.joining_date else "--",
                        self._visit_times_str(att),
                        self._clean_duration(getattr(att, "total_day_duration_str", None)),
                        getattr(att, "visits_count", len(getattr(att, "visit_list", []) or [])) or 1,
                    ]
                else:
                    m = att.member
                    gender_age = m.get_gender_display() if hasattr(m, "get_gender_display") else (m.gender or "")
                    if m.age is not None:
                        gender_age = f"{gender_age or '--'} / {m.age} yrs"
                    else:
                        gender_age = gender_age or "--"
                    values = [
                        att.date.strftime("%d %b %Y"),
                        m.member_id,
                        m.full_name,
                        m.mobile_number or "--",
                        m.membership_plan.name if m.membership_plan else "General Plan",
                        gender_age,
                        m.email or "--",
                        m.membership_start_date.strftime("%d %b %Y") if m.membership_start_date else "--",
                        m.membership_end_date.strftime("%d %b %Y") if m.membership_end_date else "--",
                        self._visit_times_str(att),
                        self._clean_duration(getattr(att, "total_day_duration_str", None)),
                        getattr(att, "visits_count", len(getattr(att, "visit_list", []) or [])) or 1,
                    ]

                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = cell_font
                    cell.border = border
                    # Entry/Exit column is multi-line; keep it left-aligned.
                    cell.alignment = left if col == entry_exit_col else center
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="F5F8FB", end_color="F5F8FB", fill_type="solid")
                row += 1
                total_records += 1

        if total_records == 0:
            ws.merge_cells(f"A{row}:{last_col}{row}")
            empty_cell = ws.cell(row=row, column=1, value="No records found for the selected filters.")
            empty_cell.font = Font(name="Calibri", size=10, italic=True, color="999999")
            empty_cell.alignment = center

        # ---- Column widths ----
        if is_trainers:
            widths = [14, 14, 24, 16, 20, 16, 26, 14, 22, 16, 9]
        else:
            widths = [14, 14, 24, 16, 20, 16, 26, 16, 16, 22, 16, 9]
        for idx, width in enumerate(widths[: len(headers)], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = f"A{header_row + 1}"

        # ---- Stream response ----
        filename = (
            f"entrix_{'trainer' if is_trainers else 'member'}_attendance_"
            f"{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        )
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
